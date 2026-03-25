from __future__ import annotations

import gzip
import json
from pathlib import Path

from openpyxl import load_workbook

import edge_workspace_links as mod


def wrap_value(value):
    return {"value": value}


def make_tab(url: str, title: str = "Title", current_index: int = 0) -> dict:
    return {
        "storage": {"currentNavigationIndex": wrap_value(current_index)},
        "subdirectories": {
            "navigationStack": {
                "subdirectories": {
                    str(current_index): {
                        "storage": {
                            "virtualUrl": wrap_value(url),
                            "title": wrap_value(title),
                        }
                    }
                }
            }
        },
    }


def make_favorite(url: str, title: str = "Favorite") -> dict:
    return {"value": {"nodeType": "1", "url": url, "title": title}}


def make_workspace_payload(
    *,
    tabs: list[tuple[str, str]] | None = None,
    favorites: list[tuple[str, str]] | None = None,
) -> dict:
    webcontents = {}
    for index, (url, title) in enumerate(tabs or []):
        webcontents[f"tab-{index}"] = make_tab(url, title, index)

    favorite_storage = {}
    for index, (url, title) in enumerate(favorites or []):
        favorite_storage[f"favorite-{index}"] = make_favorite(url, title)

    return {
        "content": {
            "subdirectories": {
                "tabstripmodel": {
                    "subdirectories": {
                        "webcontents": {
                            "subdirectories": webcontents,
                        }
                    }
                },
                "favorites": {
                    "storage": favorite_storage,
                },
            }
        }
    }


def make_edge_bytes(*payloads: dict, prefix: bytes = b"header", suffix: bytes = b"tail") -> bytes:
    data = bytearray(prefix)
    for payload in payloads:
        data.extend(gzip.compress(json.dumps(payload).encode("utf-8")))
    data.extend(suffix)
    return bytes(data)


def write_edge_file(path: Path, *payloads: dict, raw_bytes: bytes | None = None) -> Path:
    if raw_bytes is not None:
        path.write_bytes(raw_bytes)
    else:
        path.write_bytes(make_edge_bytes(*payloads))
    return path


def workbook_metrics(sheet) -> dict[str, int]:
    return {
        row[0].value: row[1].value
        for row in sheet.iter_rows(min_row=2)
        if row[0].value is not None
    }


def test_scan_gzip_payloads_deduplicates_payloads_and_extracts_workspace_data() -> None:
    first = make_workspace_payload(
        tabs=[
            ("https://example.com/a", "Alpha"),
            ("edge://settings", "Internal"),
            ("https://example.com/a", "Alpha duplicate"),
        ],
        favorites=[
            ("https://example.com/a", "Alpha favorite"),
            ("https://example.com/b", "Bravo"),
        ],
    )
    second = make_workspace_payload(
        tabs=[("https://example.com/c", "Charlie")],
        favorites=[("https://example.com/d", "Delta")],
    )

    payload_scan = mod.scan_gzip_payloads(
        make_edge_bytes(first, first, second, prefix=b"noise", suffix=b"tail")
    )
    diagnostics = mod.extract_workspace_data(payload_scan.payloads)

    assert payload_scan.had_gzip_magic is True
    assert len(payload_scan.payloads) == 2
    assert [link.url for link in diagnostics.tabs] == [
        "https://example.com/a",
        "edge://settings",
        "https://example.com/a",
        "https://example.com/c",
    ]
    assert [link.title for link in diagnostics.favorites] == [
        "Alpha favorite",
        "Bravo",
        "Delta",
    ]


def test_filter_links_excludes_explicit_schemes() -> None:
    links = [
        mod.LinkRecord(url="https://example.com", title="One"),
        mod.LinkRecord(url="chrome://settings", title="Two"),
        mod.LinkRecord(url="file:///tmp/test.txt", title="Three"),
    ]

    assert mod.filter_links(links, {"chrome", "file"}) == [
        mod.LinkRecord(url="https://example.com", title="One")
    ]


def test_main_exports_tabs_and_favorites_with_expected_dedupe(tmp_path: Path) -> None:
    write_edge_file(
        tmp_path / "sample.edge",
        make_workspace_payload(
            tabs=[
                ("https://example.com/shared", "Tab title"),
                ("https://example.com/shared", "Tab title duplicate"),
                ("https://example.com/tab-only", "Tab only"),
            ],
            favorites=[
                ("https://example.com/shared", "Favorite title"),
                ("https://example.com/favorite-only", "Favorite only"),
            ],
        ),
    )
    output_path = tmp_path / "out.xlsx"

    exit_code = mod.main(["--input", str(tmp_path), "--output", str(output_path)])

    assert exit_code == 0

    workbook = load_workbook(output_path)
    links_sheet = workbook["Links"]
    summary = workbook_metrics(workbook["Summary Report"])
    per_file_rows = list(workbook["Per File Report"].iter_rows(min_row=2, values_only=True))

    assert [links_sheet.cell(row=2, column=index).value for index in range(1, 5)] == [
        "sample.edge",
        "favorite",
        "https://example.com/shared",
        "Favorite title",
    ]
    assert links_sheet["C2"].hyperlink.target == "https://example.com/shared"
    assert [links_sheet.cell(row=3, column=index).value for index in range(1, 5)] == [
        "sample.edge",
        "favorite",
        "https://example.com/favorite-only",
        "Favorite only",
    ]
    assert [links_sheet.cell(row=4, column=index).value for index in range(1, 5)] == [
        "sample.edge",
        "tab",
        "https://example.com/tab-only",
        "Tab only",
    ]

    assert summary["workspace_files_processed"] == 1
    assert summary["extracted_tabs_total"] == 3
    assert summary["extracted_favorites_total"] == 2
    assert summary["exported_links_total"] == 3
    assert summary["unique_exported_urls"] == 3
    assert per_file_rows[0][0] == "sample.edge"
    assert per_file_rows[0][1] == "ok"
    assert per_file_rows[0][3] == 3
    assert per_file_rows[0][4] == 2
    assert per_file_rows[0][5] == 3


def test_write_output_escapes_formula_like_values_and_preserves_safe_hyperlinks(tmp_path: Path) -> None:
    output_path = tmp_path / "report.xlsx"
    rows = [
        {
            "workspace_file": "=danger.edge",
            "source": "tab",
            "url": "https://example.com",
            "title": "+SUM(A1:A2)",
        },
        {
            "workspace_file": "normal.edge",
            "source": "tab",
            "url": '=HYPERLINK("https://bad.example")',
            "title": "Safe title",
        },
    ]
    file_rows = [
        mod.FileResult(workspace_file="=danger.edge", status="ok", exported_link_count=1),
        mod.FileResult(workspace_file="normal.edge", status="ok", exported_link_count=1),
    ]

    mod.write_output(
        rows=rows,
        summary_rows=[("exported_links_total", 2)],
        file_rows=file_rows,
        output_path=output_path,
    )

    workbook = load_workbook(output_path)
    sheet = workbook["Links"]

    assert sheet["A2"].value == "'=danger.edge"
    assert sheet["D2"].value == "'+SUM(A1:A2)"
    assert sheet["C2"].value == "https://example.com"
    assert sheet["C2"].hyperlink.target == "https://example.com"
    assert sheet["C2"].data_type != "f"

    assert sheet["C3"].value == '\'=HYPERLINK("https://bad.example")'
    assert sheet["C3"].hyperlink is None
    assert sheet["C3"].data_type != "f"


def test_main_rejects_invalid_output_directory(tmp_path: Path, capsys) -> None:
    write_edge_file(
        tmp_path / "workspace.edge",
        make_workspace_payload(tabs=[("https://example.com", "Example")]),
    )

    exit_code = mod.main(
        [
            "--input",
            str(tmp_path),
            "--output",
            str(tmp_path / "missing" / "report.xlsx"),
        ]
    )

    stderr = capsys.readouterr().err
    assert exit_code == 2
    assert "Output directory not found" in stderr
    assert not (tmp_path / "missing" / "report.xlsx").exists()


def test_main_rejects_output_path_that_is_a_directory(tmp_path: Path, capsys) -> None:
    write_edge_file(
        tmp_path / "workspace.edge",
        make_workspace_payload(tabs=[("https://example.com", "Example")]),
    )
    output_dir = tmp_path / "report-dir"
    output_dir.mkdir()

    exit_code = mod.main(["--input", str(tmp_path), "--output", str(output_dir)])

    stderr = capsys.readouterr().err
    assert exit_code == 2
    assert "Output path is a directory" in stderr


def test_main_distinguishes_not_workspace_and_no_links(tmp_path: Path, capsys) -> None:
    write_edge_file(tmp_path / "empty.edge", make_workspace_payload())
    (tmp_path / "preferences.edge").write_text("not a workspace", encoding="utf-8")
    output_path = tmp_path / "report.xlsx"

    exit_code = mod.main(["--input", str(tmp_path), "--output", str(output_path)])

    stderr = capsys.readouterr().err
    assert exit_code == 0
    assert "empty.edge: no_links" in stderr
    assert "preferences.edge: not_workspace" in stderr

    workbook = load_workbook(output_path)
    metrics = workbook_metrics(workbook["Summary Report"])
    per_file_sheet = workbook["Per File Report"]
    statuses = {row[0].value: row[1].value for row in per_file_sheet.iter_rows(min_row=2)}

    assert metrics["workspace_files_processed"] == 1
    assert metrics["no_link_workspaces"] == 1
    assert metrics["not_workspace_files"] == 1
    assert metrics["exported_links_total"] == 0
    assert statuses["empty.edge"] == "no_links"
    assert statuses["preferences.edge"] == "not_workspace"


def test_mode_reports_extracted_and_exported_metrics_separately(tmp_path: Path) -> None:
    write_edge_file(
        tmp_path / "workspace.edge",
        make_workspace_payload(
            tabs=[("https://tab.example", "Tab title")],
            favorites=[("https://favorite.example", "Favorite title")],
        ),
    )
    output_path = tmp_path / "favorites.xlsx"

    exit_code = mod.main(
        [
            "--input",
            str(tmp_path),
            "--output",
            str(output_path),
            "--mode",
            "favorites",
        ]
    )

    assert exit_code == 0

    workbook = load_workbook(output_path)
    summary = workbook_metrics(workbook["Summary Report"])
    links_sheet = workbook["Links"]
    per_file_rows = list(workbook["Per File Report"].iter_rows(min_row=2, values_only=True))

    assert summary["extracted_tabs_total"] == 1
    assert summary["extracted_favorites_total"] == 1
    assert summary["exported_links_total"] == 1
    assert summary["workspace_files_with_exported_links"] == 1
    assert links_sheet.max_row == 2
    assert per_file_rows[0][3] == 1
    assert per_file_rows[0][4] == 1
    assert per_file_rows[0][5] == 1


def test_mixed_directory_exports_valid_workspaces_and_reports_read_errors(
    tmp_path: Path, monkeypatch, capsys
) -> None:
    good_path = write_edge_file(
        tmp_path / "good.edge",
        make_workspace_payload(tabs=[("https://good.example", "Good")]),
    )
    bad_path = write_edge_file(
        tmp_path / "bad.edge",
        make_workspace_payload(tabs=[("https://bad.example", "Bad")]),
    )
    output_path = tmp_path / "report.xlsx"

    original_read_bytes = mod.Path.read_bytes

    def flaky_read_bytes(self: Path) -> bytes:
        if self == bad_path:
            raise OSError("simulated read failure")
        return original_read_bytes(self)

    monkeypatch.setattr(mod.Path, "read_bytes", flaky_read_bytes)

    exit_code = mod.main(["--input", str(tmp_path), "--output", str(output_path)])

    stderr = capsys.readouterr().err
    assert exit_code == 0
    assert good_path.exists()
    assert "bad.edge: read_error" in stderr

    workbook = load_workbook(output_path)
    summary = workbook_metrics(workbook["Summary Report"])

    assert summary["workspace_files_processed"] == 1
    assert summary["read_error_files"] == 1
    assert summary["exported_links_total"] == 1


def test_corrupt_gzip_payload_is_reported_as_parse_error(tmp_path: Path, capsys) -> None:
    write_edge_file(tmp_path / "corrupt.edge", raw_bytes=mod.GZIP_MAGIC + b"broken-payload")

    exit_code = mod.main(["--input", str(tmp_path), "--output", str(tmp_path / "report.xlsx")])

    stderr = capsys.readouterr().err
    assert exit_code == 1
    assert "corrupt.edge: parse_error" in stderr
    assert "No workspace files were processed successfully." in stderr
    assert not (tmp_path / "report.xlsx").exists()


def test_oversized_payload_is_reported_as_parse_error(tmp_path: Path, monkeypatch, capsys) -> None:
    monkeypatch.setattr(mod, "MAX_PAYLOAD_BYTES", 64)
    large_title = "X" * 256
    write_edge_file(
        tmp_path / "oversized.edge",
        make_workspace_payload(tabs=[("https://example.com", large_title)]),
    )

    exit_code = mod.main(["--input", str(tmp_path), "--output", str(tmp_path / "report.xlsx")])

    stderr = capsys.readouterr().err
    assert exit_code == 1
    assert "oversized.edge: parse_error" in stderr
    assert "size guardrail" in stderr
    assert not (tmp_path / "report.xlsx").exists()
