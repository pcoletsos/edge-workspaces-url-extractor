from __future__ import annotations

import gzip
import json
from pathlib import Path

from openpyxl import load_workbook

import edge_workspace_links as mod
from tests.helpers import edge_bytes, workspace_document, write_edge_file


def workbook_metrics(sheet) -> dict[str, int]:
    return {
        row[0].value: row[1].value
        for row in sheet.iter_rows(min_row=2)
        if row[0].value is not None
    }


def test_main_exclude_internal_filters_documented_browser_urls(tmp_path: Path) -> None:
    write_edge_file(
        tmp_path / "workspace.edge",
        workspace_document(
            tabs=[
                {"url": "edge://settings", "title": "Internal"},
                {"url": "https://example.com/tab-only", "title": "Tab only"},
            ],
            favorites=[{"url": "https://example.com/favorite-only", "title": "Favorite only"}],
        ),
    )
    output_path = tmp_path / "filtered.xlsx"

    exit_code = mod.main(
        [
            "--input",
            str(tmp_path),
            "--output",
            str(output_path),
            "--exclude-internal",
            "--sort",
        ]
    )

    assert exit_code == 0

    workbook = load_workbook(output_path)
    links_sheet = workbook["Links"]
    summary = workbook_metrics(workbook["Summary Report"])

    exported_rows = list(links_sheet.iter_rows(min_row=2, values_only=True))
    assert exported_rows == [
        ("workspace.edge", "favorite", "https://example.com/favorite-only", "Favorite only"),
        ("workspace.edge", "tab", "https://example.com/tab-only", "Tab only"),
    ]
    assert summary["extracted_tabs_total"] == 2
    assert summary["extracted_favorites_total"] == 1
    assert summary["exported_links_total"] == 2


def test_main_exclude_schemes_filters_requested_schemes_in_json_output(tmp_path: Path) -> None:
    write_edge_file(
        tmp_path / "workspace.edge",
        workspace_document(
            tabs=[
                {"url": "chrome://settings", "title": "Chrome settings"},
                {"url": "file:///C:/Temp/example.txt", "title": "Local file"},
                {"url": "https://example.com/tab-only", "title": "Tab only"},
            ],
            favorites=[{"url": "https://example.com/favorite-only", "title": "Favorite only"}],
        ),
    )
    output_path = tmp_path / "filtered.json"

    exit_code = mod.main(
        [
            "--input",
            str(tmp_path),
            "--output",
            str(output_path),
            "--format",
            "json",
            "--exclude-schemes",
            "chrome",
            "file",
        ]
    )

    assert exit_code == 0

    payload = json.loads(output_path.read_text(encoding="utf-8"))
    assert [row["url"] for row in payload["links"]] == [
        "https://example.com/favorite-only",
        "https://example.com/tab-only",
    ]
    assert payload["summary"]["extracted_tabs_total"] == 3
    assert payload["summary"]["exported_links_total"] == 2


def test_main_sort_orders_links_and_file_rows_for_json_output(tmp_path: Path) -> None:
    write_edge_file(
        tmp_path / "b.edge",
        workspace_document(
            tabs=[
                {"url": "https://example.com/b-tab-b", "title": "B tab B"},
                {"url": "https://example.com/b-tab-a", "title": "B tab A"},
            ],
            favorites=[{"url": "https://example.com/b-favorite", "title": "B favorite"}],
        ),
    )
    write_edge_file(
        tmp_path / "a.edge",
        workspace_document(
            tabs=[{"url": "https://example.com/a-tab", "title": "A tab"}],
            favorites=[
                {"url": "https://example.com/a-favorite-b", "title": "A favorite B"},
                {"url": "https://example.com/a-favorite-a", "title": "A favorite A"},
            ],
        ),
    )
    output_path = tmp_path / "sorted.json"

    exit_code = mod.main(
        [
            "--input",
            str(tmp_path),
            "--output",
            str(output_path),
            "--format",
            "json",
            "--sort",
        ]
    )

    assert exit_code == 0

    payload = json.loads(output_path.read_text(encoding="utf-8"))
    assert [row["workspace_file"] for row in payload["files"]] == ["a.edge", "b.edge"]
    assert [
        (row["workspace_file"], row["source"], row["url"], row["title"]) for row in payload["links"]
    ] == [
        ("a.edge", "favorite", "https://example.com/a-favorite-a", "A favorite A"),
        ("a.edge", "favorite", "https://example.com/a-favorite-b", "A favorite B"),
        ("a.edge", "tab", "https://example.com/a-tab", "A tab"),
        ("b.edge", "favorite", "https://example.com/b-favorite", "B favorite"),
        ("b.edge", "tab", "https://example.com/b-tab-a", "B tab A"),
        ("b.edge", "tab", "https://example.com/b-tab-b", "B tab B"),
    ]


def test_process_edge_file_reports_skipped_corrupt_members_alongside_valid_data(tmp_path: Path) -> None:
    path = tmp_path / "workspace.edge"
    path.write_bytes(
        edge_bytes(
            workspace_document(
                tabs=[{"url": "https://example.com/tab-only", "title": "Tab only"}],
                favorites=[],
            ),
            prefix=b"noise",
            suffix=mod.GZIP_MAGIC + b"broken-member",
        )
    )

    file_result, rows = mod.process_edge_file(path, "both", set())

    assert file_result.status == "ok"
    assert "1 gzip member(s) were skipped." in file_result.detail
    assert [row["url"] for row in rows] == ["https://example.com/tab-only"]


def test_process_edge_file_recovers_control_character_heavy_payloads(tmp_path: Path) -> None:
    document = workspace_document(
        tabs=[{"url": "https://example.com/tab-only", "title": "Tab only"}],
        favorites=[{"url": "https://example.com/favorite-only", "title": "Favorite only"}],
    )
    noisy_payload = b"\x00\x1f" + json.dumps(document).encode("utf-8") + b"\x07\x0b"
    path = tmp_path / "workspace.edge"
    path.write_bytes(b"noise" + gzip.compress(noisy_payload) + b"tail")

    file_result, rows = mod.process_edge_file(path, "both", set())

    assert file_result.status == "ok"
    assert file_result.extracted_tab_count == 1
    assert file_result.extracted_favorite_count == 1
    assert [row["source"] for row in rows] == ["favorite", "tab"]


def test_process_edge_file_handles_large_valid_workspace_fixture(tmp_path: Path) -> None:
    payloads = []
    for payload_index in range(4):
        payloads.append(
            workspace_document(
                tabs=[
                    {
                        "url": f"https://example.com/tabs/{payload_index}/{tab_index}",
                        "title": f"Tab {payload_index}-{tab_index}",
                    }
                    for tab_index in range(150)
                ],
                favorites=[
                    {
                        "url": f"https://example.com/favorites/{payload_index}/{favorite_index}",
                        "title": f"Favorite {payload_index}-{favorite_index}",
                    }
                    for favorite_index in range(50)
                ],
            )
        )

    path = write_edge_file(tmp_path / "large.edge", *payloads)

    file_result, rows = mod.process_edge_file(path, "both", set())

    assert file_result.status == "ok"
    assert file_result.extracted_tab_count == 600
    assert file_result.extracted_favorite_count == 200
    assert file_result.exported_link_count == 800
    assert len(rows) == 800
