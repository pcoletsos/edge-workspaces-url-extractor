#!/usr/bin/env python3
from __future__ import annotations

import argparse
import gzip
import json
import os
import subprocess
import sys
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook


def workspace_document() -> dict[str, object]:
    return {
        "payload": {
            "content": {
                "subdirectories": {
                    "tabstripmodel": {
                        "subdirectories": {
                            "webcontents": {
                                "subdirectories": {
                                    "0": {
                                        "storage": {
                                            "currentNavigationIndex": {"value": 0},
                                        },
                                        "subdirectories": {
                                            "navigationStack": {
                                                "subdirectories": {
                                                    "0": {
                                                        "storage": {
                                                            "virtualUrl": {
                                                                "value": "https://example.com/tab-only"
                                                            },
                                                            "title": {"value": "Tab only"},
                                                        }
                                                    },
                                                    "1": {
                                                        "storage": {
                                                            "virtualUrl": {"value": "edge://settings"},
                                                            "title": {"value": "Internal settings"},
                                                        }
                                                    },
                                                }
                                            }
                                        },
                                    },
                                    "1": {
                                        "storage": {
                                            "currentNavigationIndex": {"value": 0},
                                        },
                                        "subdirectories": {
                                            "navigationStack": {
                                                "subdirectories": {
                                                    "0": {
                                                        "storage": {
                                                            "virtualUrl": {"value": "edge://settings"},
                                                            "title": {"value": "Internal settings"},
                                                        }
                                                    }
                                                }
                                            }
                                        },
                                    }
                                }
                            }
                        }
                    },
                    "favorites": {
                        "storage": {
                            "0": {
                                "value": {
                                    "nodeType": "1",
                                    "url": "https://example.com/favorite-only",
                                    "title": "Favorite only",
                                }
                            }
                        }
                    },
                }
            }
        }
    }


def write_edge_fixture(path: Path) -> Path:
    payload = json.dumps(workspace_document(), separators=(",", ":")).encode("utf-8")
    path.write_bytes(b"fixture" + gzip.compress(payload) + b"tail")
    return path


def build_command(args: argparse.Namespace) -> list[str]:
    if args.exe:
        return [str(Path(args.exe).resolve())]
    return [sys.executable, "-m", args.python_module]


def build_env(args: argparse.Namespace) -> dict[str, str]:
    env = os.environ.copy()
    if args.python_path:
        injected = [str(Path(path).resolve()) for path in args.python_path]
        existing = env.get("PYTHONPATH")
        if existing:
            injected.append(existing)
        env["PYTHONPATH"] = os.pathsep.join(injected)
    return env


def assert_workbook(output_path: Path) -> None:
    workbook = load_workbook(output_path)
    links = list(workbook["Links"].iter_rows(min_row=2, values_only=True))
    summary = {
        row[0].value: row[1].value
        for row in workbook["Summary Report"].iter_rows(min_row=2)
        if row[0].value is not None
    }

    if links != [
        ("workspace.edge", "favorite", "https://example.com/favorite-only", "Favorite only"),
        ("workspace.edge", "tab", "https://example.com/tab-only", "Tab only"),
    ]:
        raise AssertionError(f"Unexpected exported rows: {links!r}")

    expected_summary = {
        "workspace_files_processed": 1,
        "extracted_tabs_total": 2,
        "extracted_favorites_total": 1,
        "exported_links_total": 2,
    }
    for metric, value in expected_summary.items():
        if summary.get(metric) != value:
            raise AssertionError(f"Unexpected summary metric {metric!r}: {summary.get(metric)!r}")


def run_smoke(command: list[str], env: dict[str, str]) -> None:
    with TemporaryDirectory() as temp_dir:
        root = Path(temp_dir)
        input_dir = root / "input"
        input_dir.mkdir()
        write_edge_fixture(input_dir / "workspace.edge")
        output_path = root / "edge_workspace_links.xlsx"

        completed = subprocess.run(
            command
            + [
                "--input",
                str(input_dir),
                "--output",
                str(output_path),
                "--exclude-internal",
                "--sort",
            ],
            cwd=Path(__file__).resolve().parents[1],
            env=env,
            capture_output=True,
            text=True,
        )

        if completed.returncode != 0:
            raise AssertionError(
                "Smoke command failed:\n"
                f"command={command!r}\n"
                f"stdout={completed.stdout}\n"
                f"stderr={completed.stderr}"
            )
        if not output_path.exists():
            raise AssertionError(f"Expected output workbook was not created: {output_path}")
        if "Wrote 2 exported link(s) from 1 workspace file(s)" not in completed.stderr:
            raise AssertionError(f"Unexpected stderr output: {completed.stderr!r}")

        assert_workbook(output_path)


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run an end-to-end CLI smoke test against a built executable or importable module."
    )
    target = parser.add_mutually_exclusive_group(required=True)
    target.add_argument("--exe", help="Path to a packaged executable to smoke-test.")
    target.add_argument(
        "--python-module",
        help="Python module to invoke with `python -m`, for example edge_workspace_links.",
    )
    parser.add_argument(
        "--python-path",
        action="append",
        default=[],
        help="Additional path to prepend to PYTHONPATH before invoking --python-module.",
    )
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    run_smoke(build_command(args), build_env(args))
    target = args.exe or args.python_module
    print(f"Smoke test passed for {target}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
