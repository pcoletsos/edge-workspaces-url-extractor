#!/usr/bin/env python3
"""
Extract open tab URLs and workspace favorites from Microsoft Edge Workspace .edge files.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
import zlib
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlsplit

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None


GZIP_MAGIC = b"\x1f\x8b"
INTERNAL_SCHEMES = {"about", "chrome", "edge", "file", "microsoft-edge"}
FORMULA_PREFIXES = ("=", "+", "-", "@")
SUCCESS_STATUSES = {"ok", "no_links"}
CONTROL_CHAR_TRANSLATION = str.maketrans({chr(i): " " for i in range(32)})
NESTED_JSON_HINTS = (
    '"content"',
    '"subdirectories"',
    '"tabstripmodel"',
    '"favorites"',
    '"webcontents"',
    '"navigationStack"',
)


@dataclass(frozen=True)
class LinkRecord:
    url: str
    title: str


@dataclass(frozen=True)
class PayloadScanResult:
    had_gzip_magic: bool
    payloads: list[bytes]
    failed_members: int = 0


@dataclass(frozen=True)
class ExtractionDiagnostics:
    tabs: list[LinkRecord]
    favorites: list[LinkRecord]
    json_objects_found: int = 0
    content_objects_found: int = 0
    workspace_markers_found: int = 0


@dataclass(frozen=True)
class FileResult:
    workspace_file: str
    status: str
    detail: str = ""
    extracted_tab_count: int = 0
    extracted_favorite_count: int = 0
    exported_link_count: int = 0


def default_input_path() -> str:
    if getattr(sys, "frozen", False):
        return str(Path(sys.executable).resolve().parent)
    return "."


def iter_gzip_offsets(data: bytes) -> Iterable[int]:
    start = 0
    while True:
        idx = data.find(GZIP_MAGIC, start)
        if idx == -1:
            return
        yield idx
        start = idx + 1


def scan_gzip_payloads(data: bytes) -> PayloadScanResult:
    payloads: list[bytes] = []
    seen: set[bytes] = set()
    had_gzip_magic = False
    failed_members = 0

    for idx in iter_gzip_offsets(data):
        had_gzip_magic = True
        try:
            decompressor = zlib.decompressobj(16 + zlib.MAX_WBITS)
            out = decompressor.decompress(data[idx:])
            if not decompressor.eof or not out:
                failed_members += 1
                continue
            digest = hashlib.sha256(out).digest()
            if digest in seen:
                continue
            seen.add(digest)
            payloads.append(out)
        except zlib.error:
            failed_members += 1

    return PayloadScanResult(
        had_gzip_magic=had_gzip_magic,
        payloads=payloads,
        failed_members=failed_members,
    )


def iter_json_objects(text: str) -> Iterable[Any]:
    decoder = json.JSONDecoder()
    idx = 0
    while idx < len(text):
        next_object = text.find("{", idx)
        next_array = text.find("[", idx)
        if next_object == -1:
            idx = next_array
        elif next_array == -1:
            idx = next_object
        else:
            idx = min(next_object, next_array)
        if idx == -1:
            return
        try:
            obj, end = decoder.raw_decode(text, idx)
        except json.JSONDecodeError:
            idx += 1
            continue
        yield obj
        idx = end


def iter_content_objects(obj: Any) -> Iterable[dict[str, Any]]:
    stack: list[Any] = [obj]
    while stack:
        current = stack.pop()
        if isinstance(current, dict):
            content = current.get("content")
            if isinstance(content, dict):
                yield content
            values = list(current.values())
            stack.extend(reversed(values))
        elif isinstance(current, list):
            stack.extend(reversed(current))
        elif isinstance(current, str):
            candidate = current.strip()
            if not candidate.startswith(("{", "[")):
                continue
            if not any(hint in candidate for hint in NESTED_JSON_HINTS):
                continue
            try:
                nested = json.loads(candidate)
            except Exception:
                continue
            stack.append(nested)


def typed_value(value: Any) -> Any:
    if isinstance(value, dict) and "value" in value:
        return value.get("value")
    return value


def has_workspace_markers(content: dict[str, Any]) -> bool:
    subdirectories = content.get("subdirectories", {})
    return isinstance(subdirectories, dict) and (
        "tabstripmodel" in subdirectories or "favorites" in subdirectories
    )


def extract_tabs_from_content(content: dict[str, Any]) -> list[LinkRecord]:
    links: list[LinkRecord] = []
    webcontents = (
        content.get("subdirectories", {})
        .get("tabstripmodel", {})
        .get("subdirectories", {})
        .get("webcontents", {})
        .get("subdirectories", {})
    )
    if not isinstance(webcontents, dict):
        return links

    for tab_data in webcontents.values():
        if not isinstance(tab_data, dict):
            continue
        storage = tab_data.get("storage", {})
        current_index = typed_value(storage.get("currentNavigationIndex"))
        if current_index is None:
            continue
        nav_stack = (
            tab_data.get("subdirectories", {})
            .get("navigationStack", {})
            .get("subdirectories", {})
        )
        if not isinstance(nav_stack, dict) or not nav_stack:
            continue
        current_key = str(current_index)
        entry = nav_stack.get(current_key)
        if not entry:
            numeric_keys = [int(key) for key in nav_stack.keys() if str(key).isdigit()]
            if numeric_keys:
                entry = nav_stack.get(str(max(numeric_keys)))
        if not entry:
            continue
        entry_storage = entry.get("storage", {})
        url = ""
        for key in ("virtualUrl", "originalRequestUrl", "url"):
            value = typed_value(entry_storage.get(key))
            if isinstance(value, str) and value:
                url = value
                break
        if not url:
            continue
        title_value = typed_value(entry_storage.get("title"))
        title = title_value if isinstance(title_value, str) else ""
        links.append(LinkRecord(url=url, title=title))

    return links


def extract_favorites_from_content(content: dict[str, Any]) -> list[LinkRecord]:
    links: list[LinkRecord] = []
    favorites = content.get("subdirectories", {}).get("favorites", {})
    if not isinstance(favorites, dict):
        return links
    storage = favorites.get("storage", {})
    if not isinstance(storage, dict):
        return links

    for entry in storage.values():
        node = typed_value(entry)
        if not isinstance(node, dict):
            continue
        node_type = node.get("nodeType")
        url = node.get("url")
        if str(node_type) != "1" or not isinstance(url, str) or not url:
            continue
        title = node.get("title")
        links.append(LinkRecord(url=url, title=title if isinstance(title, str) else ""))

    return links


def clean_json_text(text: str) -> str:
    return text.translate(CONTROL_CHAR_TRANSLATION)


def extract_workspace_data(payloads: Iterable[bytes]) -> ExtractionDiagnostics:
    tabs: list[LinkRecord] = []
    favorites: list[LinkRecord] = []
    json_objects_found = 0
    content_objects_found = 0
    workspace_markers_found = 0

    for payload in payloads:
        clean = clean_json_text(payload.decode("utf-8", errors="ignore"))
        for obj in iter_json_objects(clean):
            json_objects_found += 1
            for content in iter_content_objects(obj):
                content_objects_found += 1
                if has_workspace_markers(content):
                    workspace_markers_found += 1
                tabs.extend(extract_tabs_from_content(content))
                favorites.extend(extract_favorites_from_content(content))

    return ExtractionDiagnostics(
        tabs=tabs,
        favorites=favorites,
        json_objects_found=json_objects_found,
        content_objects_found=content_objects_found,
        workspace_markers_found=workspace_markers_found,
    )


def iter_edge_files(input_path: Path) -> list[Path]:
    if input_path.is_file():
        return [input_path]
    if input_path.is_dir():
        return sorted(input_path.glob("*.edge"))
    raise FileNotFoundError(f"Input path not found: {input_path}")


def filter_links(links: list[LinkRecord], exclude_schemes: set[str]) -> list[LinkRecord]:
    if not exclude_schemes:
        return links
    filtered: list[LinkRecord] = []
    for link in links:
        scheme = link.url.split(":", 1)[0].lower() if ":" in link.url else ""
        if scheme in exclude_schemes:
            continue
        filtered.append(link)
    return filtered


def unique_by_url(links: list[LinkRecord]) -> dict[str, str]:
    url_to_title: dict[str, str] = {}
    for link in links:
        if not link.url:
            continue
        if link.url not in url_to_title or (not url_to_title[link.url] and link.title):
            url_to_title[link.url] = link.title
    return url_to_title


def build_export_rows(
    workspace_file: str,
    tabs: list[LinkRecord],
    favorites: list[LinkRecord],
    mode: str,
    exclude_schemes: set[str],
) -> list[dict[str, str]]:
    filtered_tabs = filter_links(tabs, exclude_schemes)
    filtered_favorites = filter_links(favorites, exclude_schemes)

    tab_urls = unique_by_url(filtered_tabs) if mode in {"both", "tabs"} else {}
    favorite_urls = unique_by_url(filtered_favorites) if mode in {"both", "favorites"} else {}

    combined_urls: dict[str, tuple[str, str]] = {}
    for url, title in favorite_urls.items():
        combined_urls[url] = ("favorite", title)
    for url, title in tab_urls.items():
        if url in combined_urls:
            continue
        combined_urls[url] = ("tab", title)

    return [
        {
            "workspace_file": workspace_file,
            "source": source,
            "url": url,
            "title": title,
        }
        for url, (source, title) in combined_urls.items()
    ]


def resolve_output_path(input_path: Path, output: str | None) -> Path:
    if output:
        return Path(output).expanduser()
    base_dir = input_path if input_path.is_dir() else input_path.parent
    return base_dir / "edge_workspace_links.xlsx"


def validate_output_path(output_path: Path) -> str | None:
    parent = output_path.parent if output_path.parent != Path("") else Path(".")
    if not parent.exists():
        return f"Output directory not found: {parent}"
    if not parent.is_dir():
        return f"Output path parent is not a directory: {parent}"
    if output_path.exists() and output_path.is_dir():
        return f"Output path is a directory: {output_path}"
    return None


def safe_excel_text(value: str) -> str:
    stripped = value.lstrip()
    if stripped.startswith(FORMULA_PREFIXES):
        return f"'{value}"
    return value


def safe_hyperlink_target(url: str) -> str | None:
    stripped = url.strip()
    if not stripped:
        return None
    if stripped.startswith(FORMULA_PREFIXES):
        return None
    parsed = urlsplit(stripped)
    if not parsed.scheme:
        return None
    return stripped


def process_edge_file(
    path: Path,
    mode: str,
    exclude_schemes: set[str],
) -> tuple[FileResult, list[dict[str, str]]]:
    try:
        data = path.read_bytes()
    except OSError as exc:
        return (
            FileResult(
                workspace_file=path.name,
                status="read_error",
                detail=f"Could not read file: {exc}",
            ),
            [],
        )

    try:
        payload_scan = scan_gzip_payloads(data)
        if not payload_scan.had_gzip_magic:
            return (
                FileResult(
                    workspace_file=path.name,
                    status="not_workspace",
                    detail="No gzip workspace payloads were found.",
                ),
                [],
            )
        if not payload_scan.payloads:
            return (
                FileResult(
                    workspace_file=path.name,
                    status="parse_error",
                    detail="Found gzip members but could not decompress a workspace payload.",
                ),
                [],
            )

        diagnostics = extract_workspace_data(payload_scan.payloads)
    except Exception as exc:
        return (
            FileResult(
                workspace_file=path.name,
                status="parse_error",
                detail=f"Could not parse workspace data: {exc}",
            ),
            [],
        )

    if diagnostics.json_objects_found == 0:
        return (
            FileResult(
                workspace_file=path.name,
                status="parse_error",
                detail="Decompressed payloads did not yield readable JSON objects.",
            ),
            [],
        )

    if diagnostics.workspace_markers_found == 0:
        return (
            FileResult(
                workspace_file=path.name,
                status="not_workspace",
                detail="Decoded payloads did not contain Edge Workspace content.",
            ),
            [],
        )

    export_rows = build_export_rows(
        workspace_file=path.name,
        tabs=diagnostics.tabs,
        favorites=diagnostics.favorites,
        mode=mode,
        exclude_schemes=exclude_schemes,
    )
    extracted_tab_count = len(diagnostics.tabs)
    extracted_favorite_count = len(diagnostics.favorites)
    exported_link_count = len(export_rows)

    if extracted_tab_count == 0 and extracted_favorite_count == 0:
        detail = "Workspace metadata was found, but no tabs or favorites were extracted."
        if payload_scan.failed_members:
            detail += f" {payload_scan.failed_members} gzip member(s) were skipped."
        return (
            FileResult(
                workspace_file=path.name,
                status="no_links",
                detail=detail,
                extracted_tab_count=extracted_tab_count,
                extracted_favorite_count=extracted_favorite_count,
                exported_link_count=exported_link_count,
            ),
            export_rows,
        )

    detail = "Workspace processed successfully."
    if exported_link_count == 0:
        detail = "Workspace processed successfully, but mode/filters exported no links."
    elif payload_scan.failed_members:
        detail += f" {payload_scan.failed_members} gzip member(s) were skipped."

    return (
        FileResult(
            workspace_file=path.name,
            status="ok",
            detail=detail,
            extracted_tab_count=extracted_tab_count,
            extracted_favorite_count=extracted_favorite_count,
            exported_link_count=exported_link_count,
        ),
        export_rows,
    )


def write_output(
    rows: list[dict[str, str]],
    summary_rows: list[tuple[str, int]],
    file_rows: list[FileResult],
    output_path: Path,
) -> None:
    workbook = Workbook()

    links_sheet = workbook.active
    links_sheet.title = "Links"
    links_sheet.append(["workspace_file", "source", "url", "title"])
    for row in rows:
        display_url = safe_excel_text(row["url"])
        links_sheet.append(
            [
                safe_excel_text(row["workspace_file"]),
                row["source"],
                display_url,
                safe_excel_text(row["title"]),
            ]
        )
        url_cell = links_sheet.cell(row=links_sheet.max_row, column=3)
        hyperlink_target = safe_hyperlink_target(row["url"])
        if hyperlink_target:
            url_cell.hyperlink = hyperlink_target
            url_cell.style = "Hyperlink"

    summary_sheet = workbook.create_sheet("Summary Report")
    summary_sheet.append(["metric", "value"])
    for metric, value in summary_rows:
        summary_sheet.append([metric, value])

    per_file_sheet = workbook.create_sheet("Per File Report")
    per_file_sheet.append(
        [
            "workspace_file",
            "status",
            "detail",
            "extracted_tab_count",
            "extracted_favorite_count",
            "exported_link_count",
        ]
    )
    for row in file_rows:
        per_file_sheet.append(
            [
                safe_excel_text(row.workspace_file),
                row.status,
                safe_excel_text(row.detail),
                row.extracted_tab_count,
                row.extracted_favorite_count,
                row.exported_link_count,
            ]
        )

    for sheet in (links_sheet, summary_sheet, per_file_sheet):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            column_letter = column_cells[0].column_letter
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells
            )
            sheet.column_dimensions[column_letter].width = min(max_len + 2, 80)

    workbook.save(output_path)


def build_summary_rows(edge_files: list[Path], file_rows: list[FileResult], rows: list[dict[str, str]]) -> list[tuple[str, int]]:
    successful_rows = [row for row in file_rows if row.status in SUCCESS_STATUSES]
    files_with_tabs = sum(1 for row in successful_rows if row.extracted_tab_count > 0)
    files_with_favorites = sum(1 for row in successful_rows if row.extracted_favorite_count > 0)
    files_with_extracted_links = sum(
        1 for row in successful_rows if (row.extracted_tab_count + row.extracted_favorite_count) > 0
    )
    files_with_exported_links = sum(1 for row in successful_rows if row.exported_link_count > 0)
    no_link_workspaces = sum(1 for row in file_rows if row.status == "no_links")
    read_errors = sum(1 for row in file_rows if row.status == "read_error")
    parse_errors = sum(1 for row in file_rows if row.status == "parse_error")
    not_workspace_files = sum(1 for row in file_rows if row.status == "not_workspace")
    tabs_total = sum(row.extracted_tab_count for row in successful_rows)
    favorites_total = sum(row.extracted_favorite_count for row in successful_rows)
    links_total = len(rows)
    unique_urls = len({row["url"] for row in rows})

    return [
        ("input_files_found", len(edge_files)),
        ("workspace_files_processed", len(successful_rows)),
        ("workspace_files_with_tabs_extracted", files_with_tabs),
        ("workspace_files_with_favorites_extracted", files_with_favorites),
        ("workspace_files_with_extracted_links", files_with_extracted_links),
        ("workspace_files_with_exported_links", files_with_exported_links),
        ("no_link_workspaces", no_link_workspaces),
        ("not_workspace_files", not_workspace_files),
        ("read_error_files", read_errors),
        ("parse_error_files", parse_errors),
        ("extracted_tabs_total", tabs_total),
        ("extracted_favorites_total", favorites_total),
        ("exported_links_total", links_total),
        ("unique_exported_urls", unique_urls),
    ]


def format_status_message(file_result: FileResult) -> str:
    message = f"{file_result.workspace_file}: {file_result.status}"
    if file_result.detail:
        message += f" - {file_result.detail}"
    return message


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract open tab URLs and workspace favorites from Edge Workspace .edge files."
    )
    parser.add_argument(
        "-i",
        "--input",
        default=default_input_path(),
        help=(
            "Path to a .edge file or a directory containing .edge files. "
            "Defaults to the script/exe location."
        ),
    )
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="Output .xlsx file path.",
    )
    parser.add_argument(
        "--exclude-schemes",
        nargs="*",
        default=[],
        help="URL schemes to exclude from exported links (example: edge chrome file).",
    )
    parser.add_argument(
        "--exclude-internal",
        action="store_true",
        help="Exclude internal browser URLs (about, chrome, edge, file, microsoft-edge).",
    )
    parser.add_argument(
        "--mode",
        choices=["both", "tabs", "favorites"],
        default="both",
        help="What to export to the Links sheet (default: both). Reports still show extracted tab and favorite counts.",
    )
    parser.add_argument(
        "--sort",
        action="store_true",
        help="Sort output rows by workspace file and URL.",
    )
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    if Workbook is None:
        print(
            "Missing dependency 'openpyxl'. Install with: pip install openpyxl",
            file=sys.stderr,
        )
        return 2

    args = parse_args(argv)
    input_path = Path(args.input).expanduser()

    try:
        edge_files = iter_edge_files(input_path)
    except FileNotFoundError as exc:
        print(str(exc), file=sys.stderr)
        return 2

    if not edge_files:
        print("No .edge files found in the input path.", file=sys.stderr)
        return 1

    output_path = resolve_output_path(input_path, args.output)
    output_path_error = validate_output_path(output_path)
    if output_path_error:
        print(output_path_error, file=sys.stderr)
        return 2

    exclude_schemes = {s.lower() for s in args.exclude_schemes}
    if args.exclude_internal:
        exclude_schemes.update(INTERNAL_SCHEMES)

    rows: list[dict[str, str]] = []
    file_rows: list[FileResult] = []
    for path in edge_files:
        file_result, export_rows = process_edge_file(
            path=path,
            mode=args.mode,
            exclude_schemes=exclude_schemes,
        )
        file_rows.append(file_result)
        rows.extend(export_rows)

    if args.sort:
        rows.sort(
            key=lambda row: (
                row["workspace_file"],
                0 if row["source"] == "favorite" else 1,
                row["url"],
                row["title"],
            )
        )
        file_rows = sorted(file_rows, key=lambda row: row.workspace_file)

    for file_result in file_rows:
        if file_result.status != "ok":
            print(format_status_message(file_result), file=sys.stderr)

    successful_rows = [row for row in file_rows if row.status in SUCCESS_STATUSES]
    if not successful_rows:
        print("No workspace files were processed successfully.", file=sys.stderr)
        return 1

    summary_rows = build_summary_rows(edge_files=edge_files, file_rows=file_rows, rows=rows)

    try:
        write_output(
            rows=rows,
            summary_rows=summary_rows,
            file_rows=file_rows,
            output_path=output_path,
        )
    except OSError as exc:
        print(f"Could not write workbook: {exc}", file=sys.stderr)
        return 2

    print(
        f"Wrote {len(rows)} exported link(s) from {len(successful_rows)} workspace file(s) to {output_path}",
        file=sys.stderr,
    )
    return 0


def cli() -> None:
    raise SystemExit(main(sys.argv[1:]))


if __name__ == "__main__":
    cli()
