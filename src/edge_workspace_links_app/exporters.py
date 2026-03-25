from __future__ import annotations

import csv
import json
from pathlib import Path
from urllib.parse import urlsplit

from .models import ExportRow, FileResult


FORMULA_PREFIXES = ("=", "+", "-", "@")


def coerce_export_row(row: ExportRow | dict[str, str]) -> ExportRow:
    if isinstance(row, ExportRow):
        return row
    return ExportRow(
        workspace_file=row["workspace_file"],
        source=row["source"],
        url=row["url"],
        title=row["title"],
    )


def workbook_is_available() -> bool:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return False
    return True


def safe_excel_text(value: str) -> str:
    stripped = value.lstrip()
    if stripped.startswith(FORMULA_PREFIXES):
        return f"'{value}"
    return value


def safe_hyperlink_target(url: str) -> str | None:
    stripped = url.strip()
    if not stripped:
        return None
    if stripped.startswith(FORMULA_PREFIXES):
        return None
    parsed = urlsplit(stripped)
    if not parsed.scheme:
        return None
    return stripped


def write_output(
    rows: list[ExportRow] | list[dict[str, str]],
    summary_rows: list[tuple[str, int]],
    file_rows: list[FileResult],
    output_path: Path,
) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    export_rows = [coerce_export_row(row) for row in rows]

    links_sheet = workbook.active
    links_sheet.title = "Links"
    links_sheet.append(["workspace_file", "source", "url", "title"])
    for row in export_rows:
        display_url = safe_excel_text(row.url)
        links_sheet.append(
            [
                safe_excel_text(row.workspace_file),
                row.source,
                display_url,
                safe_excel_text(row.title),
            ]
        )
        url_cell = links_sheet.cell(row=links_sheet.max_row, column=3)
        hyperlink_target = safe_hyperlink_target(row.url)
        if hyperlink_target:
            url_cell.hyperlink = hyperlink_target
            url_cell.style = "Hyperlink"

    summary_sheet = workbook.create_sheet("Summary Report")
    summary_sheet.append(["metric", "value"])
    for metric, value in summary_rows:
        summary_sheet.append([metric, value])

    per_file_sheet = workbook.create_sheet("Per File Report")
    per_file_sheet.append(
        [
            "workspace_file",
            "status",
            "detail",
            "extracted_tab_count",
            "extracted_favorite_count",
            "exported_link_count",
        ]
    )
    for row in file_rows:
        per_file_sheet.append(
            [
                safe_excel_text(row.workspace_file),
                row.status,
                safe_excel_text(row.detail),
                row.extracted_tab_count,
                row.extracted_favorite_count,
                row.exported_link_count,
            ]
        )

    for sheet in (links_sheet, summary_sheet, per_file_sheet):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            column_letter = column_cells[0].column_letter
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells
            )
            sheet.column_dimensions[column_letter].width = min(max_len + 2, 80)

    workbook.save(output_path)


def write_json_output(
    rows: list[ExportRow] | list[dict[str, str]],
    summary_rows: list[tuple[str, int]],
    file_rows: list[FileResult],
    output_path: Path,
) -> None:
    export_rows = [coerce_export_row(row) for row in rows]
    payload = {
        "links": [
            {
                "workspace_file": row.workspace_file,
                "source": row.source,
                "url": row.url,
                "title": row.title,
            }
            for row in export_rows
        ],
        "summary": {metric: value for metric, value in summary_rows},
        "files": [
            {
                "workspace_file": row.workspace_file,
                "status": row.status,
                "detail": row.detail,
                "extracted_tab_count": row.extracted_tab_count,
                "extracted_favorite_count": row.extracted_favorite_count,
                "exported_link_count": row.exported_link_count,
            }
            for row in file_rows
        ],
    }
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def write_csv_output(
    rows: list[ExportRow] | list[dict[str, str]],
    summary_rows: list[tuple[str, int]],
    file_rows: list[FileResult],
    output_paths: list[Path],
) -> None:
    links_path, summary_path, files_path = output_paths
    export_rows = [coerce_export_row(row) for row in rows]

    with links_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["workspace_file", "source", "url", "title"])
        for row in export_rows:
            writer.writerow([row.workspace_file, row.source, row.url, row.title])

    with summary_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["metric", "value"])
        writer.writerows(summary_rows)

    with files_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "workspace_file",
                "status",
                "detail",
                "extracted_tab_count",
                "extracted_favorite_count",
                "exported_link_count",
            ]
        )
        for row in file_rows:
            writer.writerow(
                [
                    row.workspace_file,
                    row.status,
                    row.detail,
                    row.extracted_tab_count,
                    row.extracted_favorite_count,
                    row.exported_link_count,
                ]
            )
